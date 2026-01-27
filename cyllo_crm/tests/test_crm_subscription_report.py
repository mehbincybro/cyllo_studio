# -*- coding: utf-8 -*-
#############################################################################
#
#    Cyllo Pvt. Ltd.
#
#    Copyright (C) 2025-TODAY Cyllo(<https://www.cyllo.com>)
#    Author: Cyllo(<https://www.cyllo.com>)
#
#    You can modify it under the terms of the GNU LESSER
#    GENERAL PUBLIC LICENSE (LGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU LESSER GENERAL PUBLIC LICENSE (LGPL v3) for more details.
#
#    You should have received a copy of the GNU LESSER GENERAL PUBLIC LICENSE
#    (LGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
#############################################################################
from odoo.tests.common import TransactionCase
from unittest.mock import patch, MagicMock
from datetime import datetime, timedelta
from psycopg2.errors import UniqueViolation
from odoo.tools import mute_logger
from odoo.addons.cyllo_crm.models.crm_subscription_report import _get_next_execution_date
from odoo.addons.cyllo_crm.models.crm_subscription_report import _get_interval
from io import BytesIO
import openpyxl


class TestCrmSubscriptionReport(TransactionCase):
    """
    This test class covers the following areas:

    1. **Helper Functions:**
        - `_get_next_execution_date`: Validates the next scheduled execution date
          based on report frequency (weekly, monthly, yearly) and default fallback.
        - `_get_interval`: Validates that the correct interval type and number
          are returned for different frequency values.

    2. **Constraints:**
        - Ensures that the `name` field of `crm.subscription.report` is unique
          by testing the SQL unique constraint.

    3. **CRUD Operations:**
        - `create`: Validates the creation of a subscription report and ensures
          that the corresponding `ir.cron` scheduled action is automatically created.
        - `write`: Tests updating a subscription report's name and frequency,
          and ensures the associated `ir.cron` is updated accordingly.

    4. **Report Generation and Email:**
        - `send_report_email`: Confirms that generating a report creates an
          attachment and triggers the email template correctly.
        - `get_xlsx_report`: Validates XLSX report generation, both for first-time
          runs (no cron) and subsequent runs (with existing cron jobs).
          Ensures that report data includes expected lead records.

    5. **Setup:**
        - Creates a test company, user, leads, and multiple subscription reports.
        - Assigns relevant `crm.lead` fields to reports for testing XLSX generation.

    This class uses `unittest.mock.patch` for deterministic behavior of date-dependent
    functions and `mute_logger` to suppress expected SQL errors during constraint testing.

    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        cls.company = cls.env['res.company'].create({
            'name': 'Test Company',
        })
        lead_fields = cls.env['ir.model.fields'].search([
            ('model', '=', 'crm.lead'),
            ('name', 'in', ['id', 'name'])
        ])

        cls.fixed_now = datetime(2025, 9, 1,
                                 10, 0, 0)
        cls.user = cls.env['res.users'].create({
            'name': 'Test User',
            'login': 'test_user',
            'company_id': cls.company.id,
            'company_ids': [(6, 0, [cls.company.id])]
        })
        cls.subscription = cls.env['crm.subscription.report'].create({
            'name': 'Initial Report',
            'frequency': 'monthly',
            'lead_fields': [(6, 0, lead_fields.ids)],
            'rule_to_apply': "[('priority', '=', '1')]",
            'user_id': cls.user.id,
        })
        cls.report = cls.env['crm.subscription.report'].create({
            'name': 'Test Report',
            'company_id': cls.company.id,
            'frequency': 'monthly',
        })
        cls.new_report = cls.env['crm.subscription.report'].create({
            'name': 'New Report',
            'frequency': 'weekly',
            'company_id': cls.company.id,
            'user_id': cls.user.id,
        })
        field_ids = cls.env['ir.model.fields'].search([
            ('model', '=', 'crm.lead'),
            ('name', 'in', ['name', 'type', 'stage_id'])
        ])
        cls.new_report.lead_fields = [(6, 0, field_ids.ids)]
        cls.lead = cls.env['crm.lead'].create({
            'name': 'Sample Lead',
            'type': 'opportunity',
            'user_id': cls.user.id,
            'company_id': cls.company.id,
        })

    def test_get_next_execution_date(self):
        """
        Test the `_get_next_execution_date` function for all frequency cases.

        This test uses `unittest.mock.patch` to mock `datetime.now()` so that
        the function behaves deterministically with a fixed reference date (`self.fixed_now`).

        It validates:
        - When frequency is "weekly", the returned date is 7 days after the fixed date.
        - When frequency is "monthly", the returned date is 30 days after the fixed date.
        - When frequency is "yearly", the returned date is 365 days after the fixed date.
        - When frequency is an invalid string (e.g., "random"), it defaults to 7 days after the fixed date.
        """
        with (patch('odoo.addons.cyllo_crm.models.crm_subscription_report.datetime') as
              mock_datetime):
            mock_datetime.now.return_value = self.fixed_now
            mock_datetime.side_effect = lambda *args, **kw: datetime(*args, **kw)

            result = _get_next_execution_date('weekly')
            expected = self.fixed_now + timedelta(days=7)
            self.assertEqual(result, expected)

        with (patch('odoo.addons.cyllo_crm.models.crm_subscription_report.datetime') as
              mock_datetime):
            mock_datetime.now.return_value = self.fixed_now
            mock_datetime.side_effect = lambda *args, **kw: datetime(*args, **kw)

            result = _get_next_execution_date('monthly')
            expected = self.fixed_now + timedelta(days=30)
            self.assertEqual(result, expected)

        with (patch('odoo.addons.cyllo_crm.models.crm_subscription_report.datetime') as
              mock_datetime):
            mock_datetime.now.return_value = self.fixed_now
            mock_datetime.side_effect = lambda *args, **kw: datetime(*args, **kw)

            result = _get_next_execution_date('yearly')
            expected = self.fixed_now + timedelta(days=365)
            self.assertEqual(result, expected)

        with patch('odoo.addons.cyllo_crm.models.crm_subscription_report.datetime') as mock_datetime:
            mock_datetime.now.return_value = self.fixed_now
            mock_datetime.side_effect = lambda *args, **kw: datetime(*args, **kw)

            result = _get_next_execution_date('random')
            expected = self.fixed_now + timedelta(days=7)
            self.assertEqual(result, expected)

    def test_get_interval(self):
        """
        Test the `_get_interval` helper function with different frequency values.
        This test validates that the function correctly returns the expected
        interval unit (e.g., weeks, months, days) and count for supported
        frequencies. It also checks the fallback case when an unknown
        frequency is provided.

        Cases covered:
            - 'weekly'  → ('weeks', 1)
            - 'monthly' → ('months', 1)
            - 'yearly'  → ('months', 12)
            - unknown frequency (e.g., 'random') → ('days', 7)
        """
        result = _get_interval('weekly')
        self.assertEqual(result, ('weeks', 1))
        result = _get_interval('monthly')
        self.assertEqual(result, ('months', 1))
        result = _get_interval('yearly')
        self.assertEqual(result, ('months', 12))
        result = _get_interval('random')
        self.assertEqual(result, ('days', 7))

    def test_unique_name_constraint(self):
        """
        Test SQL constraint to ensure name is unique
        """
        lead_fields = self.env['ir.model.fields'].search([
            ('model', '=', 'crm.lead'),
            ('name', 'in', ['id', 'name'])
        ])

        report_1 = self.env['crm.subscription.report'].create({
            'name': 'Monthly Report',
            'frequency': 'monthly',
            'lead_fields': [(6, 0, lead_fields.ids)],
        })
        self.assertTrue(report_1)
        with mute_logger('odoo.sql_db'):
            with self.assertRaises(UniqueViolation):
                self.env['crm.subscription.report'].create({
                    'name': 'Monthly Report',
                    'frequency': 'monthly',
                    'lead_fields': [(6, 0, lead_fields.ids)],
                })

    def test_create(self):
        """
        Test the creation of a CRM Subscription Report and the automatic
        creation of the corresponding scheduled action (ir.cron).

        Steps:
            1. Prepare lead fields from the crm.lead model to be used in the report.
            2. Create a crm.subscription.report record with a given frequency,
               assigned user, and the selected lead fields.
            3. Verify that an ir.cron job is automatically created for this report.
            4. Assert that:
                - The cron exists and is linked to the created report.
                - The cron's name contains the report's name.
                - The cron's user matches the report's user.
                - The cron code contains the report ID, ensuring the correct method
                  will be executed.
        """
        lead_fields = self.env['ir.model.fields'].search([
            ('model', '=', 'crm.lead'),
            ('name', 'in', ['id', 'name'])
        ])
        report = self.env['crm.subscription.report'].create({
            'name': 'Monthly SalesReport',
            'frequency': 'monthly',
            'user_id': self.user.id,
            'lead_fields': [(6, 0, lead_fields.ids)],
        })
        cron = self.env['ir.cron'].search([('ref_model_id', '=', report.id)],
                                          limit=1)
        self.assertTrue(cron)
        self.assertIn(report.name, cron.name)
        self.assertEqual(cron.user_id.id, report.user_id.id)
        self.assertIn(str(report.id), cron.code)

    def test_write(self):
        """
        Test the overridden `write` method of `crm.subscription.report`.

        This test ensures that when a subscription report's `name` or `frequency` is updated:
        1. The associated `ir.cron` record is updated accordingly.
        2. The cron job's `name` reflects the updated report name.
        3. The cron job's `interval_type` and `interval_number` reflect the updated frequency.
        4. The cron job's `code` still references the correct subscription report ID.

        The test patches `_get_next_execution_date` and `_get_interval` to control
        the returned values for predictable assertions.
        """
        cron = self.env['ir.cron'].search([('ref_model_id', '=',
                                            self.subscription.id)], limit=1)
        self.assertTrue(cron)
        with patch('odoo.addons.cyllo_crm.models.crm_subscription_report.'
                   '_get_next_execution_date') as mock_nextcall,\
             patch('odoo.addons.cyllo_crm.models.crm_subscription_report.'
                   '_get_interval') as mock_interval:
            mock_now = datetime(2025, 9, 1,
                                10, 0, 0)
            mock_nextcall.return_value = mock_now + timedelta(days=30)
            mock_interval.return_value = ('months', 1)

            self.subscription.write({
                'name': 'Updated Report',
                'frequency': 'yearly',
            })
            cron = self.env['ir.cron'].browse(cron.id)
            self.assertEqual(cron.name, f"Report - {self.subscription.name}")
            self.assertEqual(cron.interval_type, 'months')
            self.assertEqual(cron.interval_number, 1)
            self.assertIn(str(self.subscription.id), cron.code)

    def test_send_report_email(self):
        """Test that send_report_email creates attachment and
        triggers email template"""

        with patch.object(
                self.env['crm.subscription.report'].__class__,
                'get_xlsx_report',
                return_value=b'test_data'
        ) as mock_get_xlsx:
            with patch(
                    'odoo.addons.mail.models.mail_template.'
                    'MailTemplate.send_mail',
                    return_value=True) as mock_send_mail:
                self.report.send_report_email(self.report.id)

                mock_get_xlsx.assert_called_once_with(self.report.id)

                mock_send_mail.assert_called_once()

                attachment = self.env['ir.attachment'].search([
                    ('res_model', '=', 'crm.subscription.report'),
                    ('res_id', '=', self.report.id)
                ])
                self.assertTrue(attachment, "Attachment should be created")

    def test_get_xlsx_report(self):
        """
            Test the XLSX report generation for a CRM Subscription Report.

        Steps tested:
        1. Generate report when no cron exists (first run).
        2. Verify the output is bytes and not empty.
        3. Load XLSX and check that expected lead names appear.
        4. Create a cron job with `lastcall` and `nextcall`.
        5. Generate the report again and verify data is still correct.
        """
        file_data = self.new_report.get_xlsx_report(self.new_report.id)
        self.assertIsInstance(file_data, bytes)
        buffer = BytesIO(file_data)
        self.assertGreater(len(buffer.read()), 0)

        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer)
        sheet = wb.active

        all_values = [cell.value for row in sheet.iter_rows() for cell in row if
                      cell.value]

        self.assertIn('Sample Lead', all_values)

        self.env['ir.cron'].search(
            [('ref_model_id', '=', self.new_report.id)]).unlink()

        cron = self.env['ir.cron'].create({
            'name': 'Test Cron',
            'model_id': self.env['ir.model']._get('crm.subscription.report').id,
            'ref_model_id': self.new_report.id,
            'state': 'code',
            'nextcall': datetime.now() + timedelta(days=1),
            'lastcall': datetime.now() - timedelta(days=1),
        })

        file_data = self.new_report.get_xlsx_report(self.new_report.id)
        self.assertIsInstance(file_data, bytes)

        buffer = BytesIO(file_data)
        wb = openpyxl.load_workbook(buffer)
        sheet = wb.active
        all_values = [cell.value for row in sheet.iter_rows() for cell in row if
                      cell.value]

        self.assertIn('Sample Lead', all_values)
