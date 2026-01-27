# -*- coding: utf-8 -*-
#############################################################################
#
#    Cyllo Pvt. Ltd.
#
#    Copyright (C) 2025-TODAY Cyllo(<https://www.cyllo.com>)
#    Author: Cyllo(<https://www.cyllo.com>)
#
#    You can modify it under the terms of the GNU LESSER
#    GENERAL PUBLIC LICENSE (LGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU LESSER GENERAL PUBLIC LICENSE (LGPL v3) for more details.
#
#    You should have received a copy of the GNU LESSER GENERAL PUBLIC LICENSE
#    (LGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
#############################################################################
import io
from datetime import date, datetime
from unittest.mock import patch, MagicMock
from httpx import stream
from openpyxl import load_workbook
from odoo.tests.common import TransactionCase


class TestCommissionReport(TransactionCase):
    """

    """
    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        cls.company = cls.env.company
        cls.currency = cls.company.currency_id
        cls.user = cls.env['res.users'].create({
            'name': 'Dashboard User',
            'login': 'dashboard_user',
        })
        cls.team = cls.env['crm.team'].create({
            'name': 'Test Team',
        })
        cls.partner = cls.env['res.partner'].create({
            'name': 'test partner',
            'user_id': cls.user.id,
            'team_id': cls.team.id,
        })
        cls.type = cls.env['commission.type'].create({'name': 'Test plan report Type'})
        cls.plan = cls.env['commission.plan'].create({
            'name': 'Plan Alpha',
            'company_id': cls.company.id,
            'team_id': cls.team.id,
            'type_id': cls.type.id,
            'state': 'approved',
            'date_from': date.today(),
            'date_to': date.today().replace(month = 12, day = 31),
            'total_commission': 500.00,
            'sales_people_user_ids': [(6, 0, [cls.user.id])],
        })
        cls.sale_order = cls.env['sale.order'].create({
            'partner_id': cls.partner.id,
            'user_id': cls.user.id,
            'team_id': cls.team.id,
            'is_paid': True,
            'amount_untaxed': 200.00,
            'order_date': date.today(),
        })
        cls.sale_order_line = cls.env['sale.order.line'].create({
            'order_id': cls.sale_order.id,
            'product_id': cls.env.ref('product.product_product_1').id,
            'price_unit': 100.00,
            'product_uom_qty': 2,
        })
        cls.commission_report = cls.env['commission.report'].create({
            'plan_id': cls.plan.id,
            'user_id': cls.user.id,
            'company_id': cls.company.id,
            'period_name': 'Aug',
            'commission_amount': 50.00,
            'achieve_amount': 200.00,
            'order_id': cls.sale_order.id,
            'order_ids': [(6, 0, [cls.sale_order.id])],
            'orderline_ids': [(6, 0, [cls.sale_order_line.id])],
            'team_id': cls.team.id,
            'date_from': cls.plan.date_from,
            'date_to': cls.plan.date_to,
            'currency_id': cls.currency.id,
        })
        cls.lead = cls.env['crm.lead'].create({
            'name': 'Test Opportunity',
            'user_id': cls.user.id,
            'team_id': cls.team.id,
        })
        cls.won_stage = cls.env['crm.stage'].create({
            'name': 'Won',
            'is_won': True,
            'team_id': cls.team.id,
        })

        cls.won_lead = cls.env['crm.lead'].create({
            'name': 'Won Opportunity',
            'type': 'opportunity',  # required
            'user_id': cls.user.id,
            'team_id': cls.team.id,
            'company_id': cls.company.id,
            'stage_id': cls.won_stage.id,
            'active': True,
            'probability': 100,
        })
        cls.won_lead.action_set_won()
        cls.won_lead.order_ids = [(6, 0, [cls.sale_order.id])]
        cls.new_report = cls.env['commission.report'].create({
            'user_id': cls.env.user.id,
            'company_id': cls.company.id,
            'period_name': 'Test',
            'commission_amount': 100.00,
            'achieve_amount': 200.00,
            'date_from': '2025-01-01',
            'date_to': '2025-12-31',
            'currency_id': cls.currency.id,
        })

    def test_get_dashboard_data(self):
        """

        """
        report = self.commission_report
        dashboard_data = report.get_dashboard_data()
        self.assertIn('access', dashboard_data)
        self.assertEqual(dashboard_data['access'][0]['id'], self.env.user.id)
        self.assertIn('_is_manager', dashboard_data['access'][0])

        self.assertIn('plans', dashboard_data)
        self.assertTrue(any(plan['id'] == self.plan.id for plan in dashboard_data['plans']))

        self.assertIn('commissions', dashboard_data)
        self.assertTrue(any(c['id'] == report.id for c in dashboard_data['commissions']))

        self.assertEqual(dashboard_data['currency'], self.currency.name)
        self.assertEqual(dashboard_data['currency_symbol'], self.currency.symbol)

        self.assertIn('salespeople', dashboard_data)
        self.assertTrue(any(c['salesperson_id'] == self.user.id for c in dashboard_data['commissions']))

        self.assertIn('teams', dashboard_data)
        self.assertIsInstance(dashboard_data['teams'], list)
        self.assertTrue(all('id' in team and 'name' in team for team in dashboard_data['teams']))

        self.assertIn('sale_orders', dashboard_data)
        self.assertTrue(all('id' in order and 'name' in order for order in dashboard_data['sale_orders']))

        self.assertIn('orderlines', dashboard_data)
        self.assertTrue(any(line['id'] == self.sale_order_line.id for line in dashboard_data['orderlines']))

        self.assertIn('customers', dashboard_data)
        self.assertTrue(any(c['id'] == self.partner.id for c in dashboard_data['customers']))

        self.assertIn('opportunities', dashboard_data)
        self.assertTrue(any(lead['id'] == self.lead.id for lead in dashboard_data['opportunities']))

        self.assertIn('view_ids', dashboard_data)
        self.assertIn('contribution', dashboard_data['view_ids'])
        self.assertIn('target', dashboard_data['view_ids'])

    def test_get_xlsx_report(self):
        """Test XLSX generation with fake commission data"""

        report_model = self.env['commission.report']

        # 🔹 Fake report data
        fake_data = {
            "commissions": [
                {
                    "salesperson": "John Doe",
                    "plan_name": "Plan A",
                    "plan_type": "Fixed",
                    "period_name": "Aug",
                    "sale_amount": 500.0,
                    "commission_amount": 100.0,
                },
                {
                    "salesperson": "John Snow",
                    "plan_name": "Plan B",
                    "plan_type": "Variable",
                    "period_name": "Sep",
                    "sale_amount": 200.0,
                    "commission_amount": 50.0,
                },
            ],
            "filters": {"date_from": "2025-01-01", "date_to": "2025-12-31"},
            "summary": {
                "total_sales": 700.0,
                "total_commissions": 150.0,
                "total_reps": 2,
                "contributions": 2,
                "targets": 5,
            },
            "leaderboardData": [
                {"rank": 1, "name": "John Doe", "sale": 500, "commission": 100},
                {"rank": 2, "name": "John Snow", "sale": 200, "commission": 50},
            ],
        }

        response = MagicMock()
        xlsx_content = io.BytesIO()
        response.stream = xlsx_content

        report_model.get_xlsx_report(fake_data, response)

        workbook = load_workbook(io.BytesIO(xlsx_content.getvalue()))
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))

        headers = None
        for row in rows:
            if row and "Salesperson" in row:
                headers = [c for c in row if c]
                break

        expected_headers = [
            "Sl.No", "Salesperson", "Plan", "Plan Type",
            "Period", "Sale Amount", "Commission Amount"
        ]
        self.assertEqual(headers, expected_headers, "Headers do not match")

        john_doe_row = next(r for r in rows if r and "John Doe" in r)
        self.assertIn("Plan A", john_doe_row)
        self.assertIn("Fixed", john_doe_row)
        self.assertIn("Aug", john_doe_row)
        self.assertIn(500.0, john_doe_row)
        self.assertIn(100.0, john_doe_row)

        john_snow_row = next(r for r in rows if r and "John Snow" in r)
        self.assertIn("Plan B", john_snow_row)
        self.assertIn("Variable", john_snow_row)
        self.assertIn("Sep", john_snow_row)
        self.assertIn(200.0, john_snow_row)
        self.assertIn(50.0, john_snow_row)

        summary_labels = [c for row in rows for c in row if c and "Total" in str(c)]
        self.assertIn("Total Sales", summary_labels)
        self.assertIn("Total Commission", summary_labels)

        leaderboard_names = [c for row in rows for c in row if c in ["John Doe", "John Snow"]]
        self.assertIn("John Doe", leaderboard_names)
        self.assertIn("John Snow", leaderboard_names)
