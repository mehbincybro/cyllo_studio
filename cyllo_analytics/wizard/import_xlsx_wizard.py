import base64
import io
import datetime
import csv
import json
from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError:
    openpyxl = None
    column_index_from_string = None

class ImportXlsxWizardSheet(models.TransientModel):
    _name = 'import.xlsx.wizard.sheet'
    _description = 'Import XLSX Wizard Sheet'

    wizard_id = fields.Many2one('import.xlsx.wizard', string='Wizard', ondelete='cascade')
    name = fields.Char(string='Sheet Name', required=True)


class ImportXlsxWizard(models.TransientModel):
    _name = 'import.xlsx.wizard'
    _description = 'Import XLSX Dynamic Table Wizard'

    file = fields.Binary(string='File', required=True)
    file_name = fields.Char(string='File Name')
    table_name = fields.Char(string='Table Name (Technical)', help="e.g. x_my_table")
    table_label = fields.Char(string='Table Label', help="e.g. My Table")
    start_row = fields.Integer(string='Start Row', default=1, help="Row number where the table headers begin.")
    end_row = fields.Integer(string='End Row', default=0, help="Row number where the data ends. Leave 0 to read all below start row.")
    start_col = fields.Char(string='Start Column', default='A', help="Column Letter (e.g. A)")
    end_col = fields.Char(string='End Column', help="Column Letter (e.g. D). Leave empty for all.")
    state = fields.Selection([('init', 'Init'), ('sheet', 'Select Sheet'), ('columns', 'Columns'), ('confirm', 'Confirm')], default='init')
    confirm_append = fields.Boolean(default=False)
    
    @api.onchange('table_label')
    def _onchange_table_label(self):
        for rec in self:
            if rec.table_label:
                # Keep alphanumeric and spaces, then replace spaces with underscores
                safe = ''.join(e for e in rec.table_label.lower() if e.isalnum() or e == ' ')
                safe_name = safe.strip().replace(' ', '_')
                
                # Check if starts with x_ or contains numbers at start. Odoo strict:
                import re
                safe_name = re.sub(r'^[^a-zA-Z]+', '', safe_name)
                
                if safe_name and not safe_name.startswith('x_'):
                    safe_name = f"x_{safe_name}"
                rec.table_name = safe_name
    sheet_id = fields.Many2one('import.xlsx.wizard.sheet', string='Select Sheet')
    column_lines = fields.One2many('import.xlsx.wizard.line', 'wizard_id', string='Columns')

    def action_analyze_file(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Please upload a file."))
        if not self.file_name:
            raise UserError(_("File name is missing. Please re-upload."))

        if self.file_name.lower().endswith(('.csv', '.json')):
            return self.action_read_first_row()

        if not openpyxl:
            raise UserError(_("Please install openpyxl python library."))
            
        file_content = base64.b64decode(self.file)
        if not file_content:
            return

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            raise UserError(_("Failed to read the excel file. Error: %s" % e))

        existing_sheets = self.env['import.xlsx.wizard.sheet'].search([('wizard_id', '=', self.id)])
        if existing_sheets:
            existing_sheets.unlink()
            
        self.write({'sheet_id': False}) 

        new_sheets = [{'wizard_id': self.id, 'name': name} for name in sheet_names]
        if new_sheets:
            self.env['import.xlsx.wizard.sheet'].create(new_sheets)

        all_sheets = self.env['import.xlsx.wizard.sheet'].search([('wizard_id', '=', self.id)])
        
        if len(all_sheets) > 1:
            self.write({'state': 'sheet', 'sheet_id': False})
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'import.xlsx.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
                'domain': [('id', 'in', all_sheets.ids)]
            }
        else:
            self.write({'sheet_id': all_sheets[0].id if all_sheets else False})
            return self.action_read_first_row()

    def action_read_first_row(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Please upload a file."))
            
        file_content = base64.b64decode(self.file)
        if not file_content:
            self.state = 'init'
            return

        first_row = []
        second_row = []
        file_ext = self.file_name.lower() if self.file_name else ''
        
        try:
            if file_ext.endswith('.csv'):
                decoded_file = file_content.decode('utf-8')
                reader = csv.reader(io.StringIO(decoded_file))
                for i, row in enumerate(reader):
                    if i == 0:
                        first_row = list(row)
                    elif i <= 100:
                        second_row.append(list(row))
                    else:
                        break
            elif file_ext.endswith('.json'):
                decoded_file = file_content.decode('utf-8')
                data = json.loads(decoded_file)
                if not isinstance(data, dict) or not data:
                    raise UserError("JSON file must contain a list of dictionaries/objects.")
                first_row = list(data[0].keys())
                for i, row_dict in enumerate(data):
                    if i > 100: break
                    mapped_row = [row_dict.get(col, '') for col in first_row]
                    second_row.append(mapped_row)
            else:
                if not openpyxl:
                    raise UserError(_("Please install openpyxl python library."))
                if not self.sheet_id:
                    raise UserError(_("Please select a sheet first."))
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                if self.sheet_id.name in wb.sheetnames:
                    sheet = wb[self.sheet_id.name]
                else:
                    sheet = wb.active
                min_r = self.start_row if self.start_row > 0 else 1
                max_r = self.end_row if self.end_row > 0 else None
                min_c = 1
                max_c = None
                
                if self.start_col and column_index_from_string:
                    try:
                        min_c = column_index_from_string(self.start_col.upper())
                    except Exception:
                        pass
                        
                if self.end_col and column_index_from_string:
                    try:
                        max_c = column_index_from_string(self.end_col.upper())
                    except Exception:
                        pass

                iterator = sheet.iter_rows(
                    min_row=min_r, 
                    max_row=max_r, 
                    min_col=min_c,
                    max_col=max_c,
                    values_only=True
                )
                for i, row in enumerate(iterator):
                    if i == 0:
                        first_row = list(row)
                    elif i <= 100:
                        second_row.append(list(row))
                    else:
                        break
        except Exception as e:
            raise UserError(_("Failed to read the file. Error: %s" % e))
            
        # Clean existing lines
        self.column_lines.unlink()
        
        lines = []
        for index, header in enumerate(first_row):
            if not header:
                header = f"Column_{index}"
            
            # Extract column values for this index
            col_values = [row[index] for row in second_row if len(row) > index and row[index] is not None and str(row[index]).strip()]
            
            field_type = 'char'
            selection_options = False
            
            if col_values:
                sample_val = col_values[0]
                
                if isinstance(sample_val, int):
                    field_type = 'integer'
                elif isinstance(sample_val, float):
                    field_type = 'float'
                elif isinstance(sample_val, bool):
                    field_type = 'boolean'
                elif isinstance(sample_val, (datetime.date, datetime.datetime)):
                    field_type = 'date'
                else:
                    # check if we can make it a selection
                    distinct_vals = set(str(v).strip() for v in col_values)
                    # Skip selection if any value contains a comma to avoid splitting issues
                    has_comma = any(',' in v for v in distinct_vals)
                    # If few distinct values compared to total sample, it's likely a selection
                    if 0 < len(distinct_vals) <= 15 and len(col_values) >= len(distinct_vals) and not has_comma:
                        field_type = 'selection'
                        selection_options = ','.join(distinct_vals)
                    
            safe_name = ''.join(e for e in str(header) if e.isalnum() or e == '_').lower()
            if not safe_name.startswith('x_'):
                safe_name = f"x_{safe_name}"

            lines.append((0, 0, {
                'column_name': str(header),
                'field_name': safe_name,
                'field_type': field_type,
                'selection_options': selection_options,
            }))
            
        self.write({
            'state': 'columns',
            'column_lines': lines,
            'table_name': self.table_name or '',
            'table_label': self.table_label or '',
            'start_row': self.start_row,
            'end_row': self.end_row,
            'start_col': self.start_col or '',
            'end_col': self.end_col or ''
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'import.xlsx.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_import_data_confirm(self):
        self.ensure_one()
        self.confirm_append = True
        return self.action_import_data()

    def action_import_data(self):
        self.ensure_one()
        
        if not self.table_name or not self.table_name.startswith('x_'):
            raise UserError(_("Table name must start with 'x_'. e.g. 'x_my_table'"))
            
        if not self.table_label:
            raise UserError(_("Please provide a Table Label."))
            
        if not self.column_lines:
            raise UserError(_("No columns defined to import."))

        # Validate manually-set selection fields have options populated
        for line in self.column_lines:
            if line.field_type == 'selection' and not line.selection_options:
                raise UserError(_(
                    "Column '%s' is set to 'Selection' but has no valid options. "
                    "This usually means its values contain commas, which are not "
                    "allowed in Odoo selection fields.\n\n"
                    "Please change the Field Type to 'Text' instead."
                ) % line.column_name)

        # Create ir.model
        IrModel = self.env['ir.model'].sudo()
        IrModelField = self.env['ir.model.fields'].sudo()
        
        existing_model = IrModel.search([('model', '=', self.table_name)])
        
        if existing_model and not self.confirm_append:
            self.write({'state': 'confirm'})
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'import.xlsx.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }
            
        if existing_model:
            model_id = existing_model
        else:
            description = self.table_label
            model_id = IrModel.create({
                'name': description,
                'model': self.table_name,
                'state': 'manual',
            })
            
        field_mapping = {}
        
        # Create fields
        for index, line in enumerate(self.column_lines):
            field_mapping[index] = line.field_name
            existing_field = IrModelField.search([
                ('model_id', '=', model_id.id),
                ('name', '=', line.field_name)
            ])
            if not existing_field:
                field_vals = {
                    'name': line.field_name,
                    'field_description': line.column_name,
                    'model_id': model_id.id,
                    'ttype': line.field_type,
                    'state': 'manual',
                }
                if line.field_type == 'selection' and line.selection_options:
                    # Convert comma-separated string to list of tuples properly
                    opts = [opt.strip() for opt in line.selection_options.split(',') if opt.strip()]
                    selection_str = str([(opt, opt) for opt in opts])
                    field_vals['selection'] = selection_str
                elif line.field_type == 'many2one':
                    if not line.relation_model_id:
                        raise UserError(_("Relation Model is required for Many2one field '%s'." % line.column_name))
                    field_vals['relation'] = line.relation_model_id.model
                
                IrModelField.create(field_vals)
                
        # Give access to ir.model.access
        access_name = f'access_{self.table_name.replace(".", "_")}'
        existing_access = self.env['ir.model.access'].sudo().search([
            ('name', '=', access_name),
            ('model_id', '=', model_id.id)
        ])
        if not existing_access:
            self.env['ir.model.access'].sudo().create({
                'name': access_name,
                'model_id': model_id.id,
                'group_id': self.env.ref('base.group_user').id,
                'perm_read': True,
                'perm_write': True,
                'perm_create': True,
                'perm_unlink': True,
            })
        
        # Important: Flush to DB and reload registry to recognize the new model/fields
        self.env.flush_all()
        self.env.registry.setup_models(self.env.cr)
                
        # Now import the data
        file_content = base64.b64decode(self.file)
        file_ext = self.file_name.lower() if self.file_name else ''
        DynamicModel = self.env[self.table_name].sudo()

        m2o_caches = {}
        for index, line in enumerate(self.column_lines):
            if line.field_type == 'many2one' and line.relation_model_id:
                m2o_caches[index] = {'model': self.env[line.relation_model_id.model].sudo(), 'cache': {}}

        vals_list = []
        if file_ext.endswith('.csv'):
            decoded_file = file_content.decode('utf-8')
            reader = csv.reader(io.StringIO(decoded_file))
            iterator = list(reader)[1:]  # Skip headers
        elif file_ext.endswith('.json'):
            data = json.loads(file_content.decode('utf-8'))
            headers = list(data[0].keys())
            iterator = [[row_dict.get(col, '') for col in headers] for row_dict in data]
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            if self.sheet_id and self.sheet_id.name in wb.sheetnames:
                sheet = wb[self.sheet_id.name]
            else:
                sheet = wb.active
                
            min_r = self.start_row if self.start_row > 0 else 1
            max_r = self.end_row if self.end_row > 0 else None
            min_c = 1
            max_c = None
            
            if self.start_col and column_index_from_string:
                try:
                    min_c = column_index_from_string(self.start_col.upper())
                except Exception:
                    pass
                    
            if self.end_col and column_index_from_string:
                try:
                    max_c = column_index_from_string(self.end_col.upper())
                except Exception:
                    pass

            rows = list(sheet.iter_rows(
                min_row=min_r, 
                max_row=max_r, 
                min_col=min_c,
                max_col=max_c,
                values_only=True
            ))
            iterator = rows[1:] if rows else []

        for i, row in enumerate(iterator):
            
            vals = {}
            for index, val in enumerate(row):
                if index in field_mapping:
                    field_name = field_mapping[index]
                    field_type_line = self.column_lines[index]
                    
                    if val is None or str(val).strip() == '':
                        continue
                        
                    # basic conversion
                    try:
                        if field_type_line.field_type == 'integer':
                            vals[field_name] = int(val)
                        elif field_type_line.field_type == 'float':
                            vals[field_name] = float(val)
                        elif field_type_line.field_type == 'boolean':
                            vals[field_name] = bool(val)
                        elif field_type_line.field_type == 'date':
                            if isinstance(val, (datetime.date, datetime.datetime)):
                                vals[field_name] = val.strftime('%Y-%m-%d')
                            else:
                                vals[field_name] = str(val) # Let Odoo try
                        elif field_type_line.field_type == 'selection':
                            val_str = str(val).strip()
                            vals[field_name] = val_str
                        elif field_type_line.field_type == 'many2one':
                            val_str = str(val).strip()
                            if not val_str:
                                continue
                            m2o_info = m2o_caches.get(index)
                            if m2o_info:
                                if val_str in m2o_info['cache']:
                                    if m2o_info['cache'][val_str]:
                                        vals[field_name] = m2o_info['cache'][val_str]
                                else:
                                    RelationModel = m2o_info['model']
                                    target_rec = False
                                    
                                    # First, if the user put an integer, try finding it by Database ID
                                    if val_str.isdigit():
                                        id_match = RelationModel.search([('id', '=', int(val_str))], limit=1)
                                        if id_match:
                                            # Mock a name_search style tuple (id, name)
                                            target_rec = [(id_match.id, id_match.display_name)]
                                            
                                    # Fallback to standard Text Name lookup
                                    if not target_rec:
                                        target_rec = RelationModel.name_search(val_str, operator='ilike', limit=1)
                                        
                                    if target_rec:
                                        rec_id = target_rec[0][0]
                                        m2o_info['cache'][val_str] = rec_id
                                        vals[field_name] = rec_id
                                    else:
                                        m2o_info['cache'][val_str] = False
                        else:
                            vals[field_name] = str(val)
                    except Exception:
                        pass # Ignore parsing errors simply for robust import
            
            if not vals.get('x_name'):
                vals['x_name'] = f"{self.table_label} ({i})"
                
            if vals:
                vals_list.append(vals)
                
        if vals_list:
            DynamicModel.create(vals_list)

        # Create Action
        action = self.env['ir.actions.act_window'].sudo().search([('res_model', '=', self.table_name)], limit=1)
        if not action:
            action = self.env['ir.actions.act_window'].sudo().create({
                'name': self.table_label,
                'res_model': self.table_name,
                'view_mode': 'tree,form',
            })
            
        # Create Registry Record
        registry_rec = self.env['custom.imported.table'].sudo().search([('model_id', '=', model_id.id)], limit=1)
        if not registry_rec:
            self.env['custom.imported.table'].sudo().create({
                'name': self.table_label,
                'model_id': model_id.id,
                'action_id': action.id,
            })

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }

class ImportXlsxWizardLine(models.TransientModel):
    _name = 'import.xlsx.wizard.line'
    _description = 'Import XLSX Wizard Line'

    wizard_id = fields.Many2one('import.xlsx.wizard', string='Wizard')
    column_name = fields.Char(string='Excel Column Name', required=True)
    field_name = fields.Char(string='Technical Field Name', required=True)
    field_type = fields.Selection([
        ('char', 'Text'),
        ('integer', 'Integer'),
        ('float', 'Float'),
        ('boolean', 'Boolean'),
        ('date', 'Date'),
        ('text', 'Long Text'),
        ('selection', 'Selection'),
        ('many2one', 'Many2one')
    ], string='Field Type', required=True, default='char')
    relation_model_id = fields.Many2one('ir.model', string='Relation Model', help="Required for Many2one fields")
    selection_options = fields.Char(string='Selection Options', help="Comma-separated options for Selection fields")
